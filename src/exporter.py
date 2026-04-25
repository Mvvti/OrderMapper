import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.utils import sort_text, sort_number, sort_facility


def sort_final_records(final_records):
    return sorted(
        final_records,
        key=lambda r: (
            sort_text(r.get("zamowienie_nr")),
            sort_facility(r.get("obiekt_id")),
            sort_number(r.get("artykul_pozycja_na_zamowieniu")),
        ),
    )


def print_sorting_preview(final_records):
    print("\n=== PIERWSZE 10 REKORDÓW PO SORTOWANIU ===")
    if final_records:
        sorted_preview_columns = [
            "zamowienie_nr",
            "oddzial_id",
            "oddzial_nazwa",
            "obiekt_id",
            "obiekt_id_dostawcy",
            "Nazwa_miejsca_dostaw.",
            "artykul_pozycja_na_zamowieniu",
        ]
        print(pd.DataFrame(final_records[:10])[sorted_preview_columns].to_string(index=False))
    else:
        print("Brak rekordów po sortowaniu.")

    print("\n=== PIERWSZE 30 REKORDÓW PO SORTOWANIU ===")
    if final_records:
        sorted_preview_columns_30 = [
            "zamowienie_nr",
            "oddzial_id",
            "obiekt_id",
            "artykul_pozycja_na_zamowieniu",
        ]
        print(pd.DataFrame(final_records[:30])[sorted_preview_columns_30].to_string(index=False))
    else:
        print("Brak rekordów po sortowaniu.")


def export_to_excel(template_file_path, template_first_sheet_name, template_columns, final_records, records, output_dir="output"):
    output_file_path = os.path.join(output_dir, "wynik_template.xlsx")
    report_file_path = os.path.join(output_dir, "raport_brakow.txt")

    os.makedirs(output_dir, exist_ok=True)

    # Bezpieczny eksport: nowy workbook, zwykłe komórki (bez tabel Excela).
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = template_first_sheet_name

    # Nagłówki jako zwykłe komórki.
    for col_index, col_name in enumerate(template_columns, start=1):
        worksheet.cell(row=1, column=col_index, value=col_name)

    # Styl nagłówka.
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", start_color="E9EEF5", end_color="E9EEF5")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_index in range(1, len(template_columns) + 1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    worksheet.row_dimensions[1].height = 34

    start_row_to_write = 2
    end_row_to_write = start_row_to_write + len(final_records) - 1

    for row_offset, final_record in enumerate(final_records):
        excel_row = start_row_to_write + row_offset
        for col_index, col_name in enumerate(template_columns, start=1):
            cell_value = final_record.get(col_name)
            if isinstance(cell_value, str):
                cell_value = cell_value.strip()
            if cell_value is None or (pd.isna(cell_value) if not isinstance(cell_value, str) else False):
                cell_value = ""
            worksheet.cell(row=excel_row, column=col_index, value=cell_value)

    # Ułatwienia arkusza.
    worksheet.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(template_columns))
    worksheet.auto_filter.ref = f"A1:{last_col_letter}1"

    # Formatowanie kolumn po nazwach.
    col_idx = {col_name: idx + 1 for idx, col_name in enumerate(template_columns)}

    date_columns = ["data_utworzenia", "termin_dostawy"]
    decimal_columns = ["artykul_cena_netto", "art.wartosc_netto"]
    int_columns = ["artykul_ilosc"]
    wrap_columns = ["artykul_nazwa", "obiekt_adres_ulica"]
    object_columns = ["obiekt_id", "obiekt_id_dostawcy", "Nazwa_miejsca_dostaw."]

    # Szerokości kolumn.
    width_map = {
        "artykul_nazwa": 52,
        "obiekt_adres_ulica": 32,
        "obiekt_adres_miasto": 20,
        "obiekt_adres_kod_pocztowy": 14,
        "data_utworzenia": 13,
        "termin_dostawy": 13,
        "artykul_ilosc": 10,
        "artykul_cena_netto": 14,
        "art.wartosc_netto": 14,
    }
    default_width = 18
    for col_name, col_number in col_idx.items():
        width = width_map.get(col_name, default_width)
        worksheet.column_dimensions[get_column_letter(col_number)].width = width

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_row_fill = PatternFill(fill_type="solid", start_color="FAFBFD", end_color="FAFBFD")
    missing_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
    discontinued_fill = PatternFill(fill_type="solid", start_color="FEE2E2", end_color="FEE2E2")

    if end_row_to_write >= start_row_to_write:
        for row_num in range(start_row_to_write, end_row_to_write + 1):
            for col_name in date_columns:
                if col_name in col_idx:
                    cell = worksheet.cell(row=row_num, column=col_idx[col_name])
                    if isinstance(cell.value, str) and cell.value.strip() != "":
                        parsed_date = pd.to_datetime(cell.value, errors="coerce")
                        if pd.notna(parsed_date):
                            cell.value = parsed_date.date()
                    if cell.value not in (None, ""):
                        cell.number_format = "yyyy-mm-dd"

            for col_name in decimal_columns:
                if col_name in col_idx:
                    cell = worksheet.cell(row=row_num, column=col_idx[col_name])
                    if cell.value not in (None, ""):
                        parsed_num = pd.to_numeric(cell.value, errors="coerce")
                        if pd.notna(parsed_num):
                            cell.value = float(parsed_num)
                            cell.number_format = "0.00"

            for col_name in int_columns:
                if col_name in col_idx:
                    cell = worksheet.cell(row=row_num, column=col_idx[col_name])
                    if cell.value not in (None, ""):
                        parsed_num = pd.to_numeric(cell.value, errors="coerce")
                        if pd.notna(parsed_num):
                            cell.value = int(parsed_num)
                            cell.number_format = "0"

            # Zawijanie tekstu dla wybranych kolumn.
            for col_name in wrap_columns:
                if col_name in col_idx:
                    cell = worksheet.cell(row=row_num, column=col_idx[col_name])
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Delikatne obramowanie + naprzemienne tło wierszy.
            for col_number in range(1, len(template_columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_number)
                cell.border = thin_border
                if row_num % 2 == 0 and cell.fill.fill_type is None:
                    cell.fill = alt_row_fill

            # Wyróżnienie braków.
            if "artykul_cena_netto" in col_idx:
                price_cell = worksheet.cell(row=row_num, column=col_idx["artykul_cena_netto"])
                if price_cell.value in (None, ""):
                    price_cell.fill = missing_fill
            for col_name in object_columns:
                if col_name in col_idx:
                    obj_cell = worksheet.cell(row=row_num, column=col_idx[col_name])
                    if obj_cell.value in (None, ""):
                        obj_cell.fill = missing_fill

            # Wyróżnienie wycofanych produktów — czerwone tło całego wiersza.
            row_offset = row_num - start_row_to_write
            if row_offset < len(final_records) and final_records[row_offset].get("_is_discontinued"):
                for col_number in range(1, len(template_columns) + 1):
                    worksheet.cell(row=row_num, column=col_number).fill = discontinued_fill

    preview_rows = []
    preview_count = min(5, len(final_records))
    preview_columns = ["oddzial_id", "oddzial_nazwa", "obiekt_id", "obiekt_id_dostawcy", "Nazwa_miejsca_dostaw."]

    preview_column_indexes = []
    for preview_col in preview_columns:
        if preview_col in template_columns:
            preview_column_indexes.append((preview_col, template_columns.index(preview_col) + 1))

    for i in range(preview_count):
        row_num = start_row_to_write + i
        row_preview = {}
        for col_name, col_index in preview_column_indexes:
            row_preview[col_name] = worksheet.cell(row=row_num, column=col_index).value
        preview_rows.append(row_preview)

    discontinued_codes = {
        str(r.get("parsed_product_code")).strip()
        for r in records
        if r.get("is_discontinued")
    }

    records_without_price = sum(1 for r in records if not r.get("match_found"))
    records_without_facility = sum(1 for r in records if not r.get("facility_match_found"))
    unmatched_product_codes = sorted(
        {
            str(r.get("parsed_product_code")).strip()
            for r in records
            if not r.get("match_found") and r.get("parsed_product_code") not in (None, "")
        }
    )
    unmatched_facilities = sorted(
        {
            str(r.get("placowka")).strip()
            for r in records
            if not r.get("facility_match_found") and r.get("placowka") not in (None, "")
        }
    )

    workbook.save(output_file_path)

    with open(report_file_path, "w", encoding="utf-8") as report_file:
        report_file.write("RAPORT BRAKÓW\n")
        report_file.write(f"Liczba rekordów bez dopasowanej ceny: {records_without_price}\n")
        report_file.write(f"Liczba rekordów bez dopasowanej placówki: {records_without_facility}\n\n")
        report_file.write("Unikalne niedopasowane kody produktów:\n")
        for code in unmatched_product_codes:
            note = " (artykuł wycofany ze sprzedaży)" if code in discontinued_codes else ""
            report_file.write(f"- {code}{note}\n")
        report_file.write("\nUnikalne niedopasowane placówki:\n")
        for facility in unmatched_facilities:
            report_file.write(f"- {facility}\n")

    print("\n=== ZAPIS PLIKU WYNIKOWEGO ===")
    print(f"Nazwa arkusza: {worksheet.title}")
    print(f"Numer wiersza startowego zapisu: {start_row_to_write}")
    print(f"Ścieżka zapisanego pliku: {output_file_path}")
    print(f"Ścieżka raportu braków: {report_file_path}")
    print(f"Liczba zapisanych rekordów: {len(final_records)}")
    print(f"Liczba rekordów bez dopasowanej ceny: {records_without_price}")
    print(f"Liczba rekordów bez dopasowanej placówki: {records_without_facility}")
    print("\nPierwsze 5 zapisanych rekordów (wybrane kolumny):")
    if preview_rows:
        print(pd.DataFrame(preview_rows, columns=preview_columns).to_string(index=False))
    else:
        print("Brak zapisanych rekordów do podglądu.")

    return {
        "output_file_path": output_file_path,
        "report_file_path": report_file_path,
        "records_count": len(final_records),
        "records_without_price": records_without_price,
        "records_without_facility": records_without_facility,
    }
