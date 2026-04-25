from __future__ import annotations

import os
from pathlib import Path
import sys as _sys
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.pdf_pipeline.pdf_parser import parse_pdf_order


_HEADER_ALIASES: dict[str, str] = {
    "Numer materiału dostawcy": "Numer_materialu_dostawcy",
    "art.wartosc_netto": "art_wartosc_netto",
    "Nazwa_miejsca_dostaw.": "Nazwa_miejsca_dostaw",
}


def _get_resource_path(rel):
    if hasattr(_sys, '_MEIPASS'):
        return Path(_sys._MEIPASS) / rel
    return Path(__file__).resolve().parents[2] / rel


def run_pdf_pipeline(pdf_paths: list[str], output_folder: str, log_fn) -> dict[str, Any]:
    def log(message: str) -> None:
        if log_fn:
            log_fn(message)

    pdf_paths = [str(path) for path in (pdf_paths or []) if str(path).strip()]

    log("PDF: start pipeline...")
    log(f"PDF: liczba plików PDF: {len(pdf_paths)}")

    log("PDF: odczyt nagłówków templatePDF.xlsx...")
    template_path = _get_resource_path("workflow/templatePDF.xlsx")
    template_sheet_name, template_headers = _load_template_headers(template_path)

    log("PDF: budowanie rekordów...")
    rows: list[dict[str, Any]] = []
    total_files = len(pdf_paths)
    for index, pdf_path in enumerate(pdf_paths, start=1):
        file_name = Path(pdf_path).name
        log(f"PDF: przetwarzanie {index}/{total_files} — {file_name}")
        parsed = parse_pdf_order(pdf_path)
        positions = parsed.get("pozycje") or []
        rows.extend(_build_row(parsed, position) for position in positions)

    log("PDF: zapis do Excela...")
    output_dir_path = Path(output_folder)
    output_dir_path.mkdir(parents=True, exist_ok=True)
    output_path = output_dir_path / "wynik_pdf_zbiorczy.xlsx"
    _write_output_file(output_path, template_sheet_name, template_headers, rows)

    log(f"PDF: zapisano plik {output_path}")

    return {
        "output_file_path": str(output_path),
        "records_count": len(rows),
    }


def _build_row(parsed: dict[str, Any], position: dict[str, Any]) -> dict[str, Any]:
    lp = position.get("lp")
    phone = str(parsed.get("zamawiajacy_telefon") or "")

    return {
        "zamowienie_nr": str(parsed.get("zamowienie_nr") or ""),
        "oddzial_id": "631-23-63-833",
        "oddzial_nazwa": "PDF",
        "obiekt_id": str(parsed.get("obiekt_id") or ""),
        "obiekt_adres_ulica": str(parsed.get("obiekt_adres_ulica") or ""),
        "obiekt_adres_miasto": str(parsed.get("obiekt_adres_miasto") or ""),
        "obiekt_adres_kod_pocztowy": str(parsed.get("obiekt_adres_kod_pocztowy") or ""),
        "zamawiajacy_imie_nazwisko": str(parsed.get("zamawiajacy_imie_nazwisko") or ""),
        "zamawiajacy_telefon": phone,
        "zamawiajacy_email": "",
        "data_utworzenia": str(parsed.get("data_utworzenia") or ""),
        "termin_dostawy": "",
        "godziny_dostawy": "",
        "odbierajacy_telefon": phone,
        "artykul_pozycja_na_zamowieniu": lp if lp is not None else "",
        "Numer_materialu_dostawcy": str(position.get("kod_odbiorcy") or ""),
        "artykul_nr_impel": "",
        "artykul_nazwa": str(position.get("nazwa") or ""),
        "artykul_ilosc": position.get("ilosc") if position.get("ilosc") is not None else "",
        "artykul_jednostka": "SZT",
        "artykul_cena_netto": position.get("cena_netto") if position.get("cena_netto") is not None else "",
        "art_wartosc_netto": round((position.get("ilosc") or 0) * (position.get("cena_netto") or 0), 3),
        "artykul_waluta": "PLN",
        "obiekt_id_dostawcy": "",
        "Nazwa_miejsca_dostaw": "",
        "Odbierajacy_towar": "",
        "Row_ID": lp if lp is not None else "",
        "Status": "",
    }


def _load_template_headers(template_path: Path) -> tuple[str, list[str]]:
    workbook = load_workbook(template_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [
        str(sheet.cell(row=1, column=column_index).value or "")
        for column_index in range(1, sheet.max_column + 1)
    ]
    workbook.close()
    return sheet.title, headers


def _write_output_file(output_path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)

    header_fill = PatternFill("solid", start_color="2A95CF", end_color="2A95CF")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, cell in enumerate(worksheet[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row_data in enumerate(rows, start=2):
        for column_index, header in enumerate(headers, start=1):
            key = _HEADER_ALIASES.get(header, header)
            worksheet.cell(row=row_index, column=column_index, value=row_data.get(key, ""))

    for col_cells in worksheet.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        worksheet.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)

    worksheet.row_dimensions[1].height = 40

    workbook.save(output_path)
